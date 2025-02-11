import requests
from model.currency_model import Currency
from openpyxl import load_workbook

def fetch_currency_data():
    # Define the API endpoint and your API key
    api_url = "https://financialmodelingprep.com/api/v3/quotes/forex"
    api_key = ""
    currencies = []

    # Define the ordered currency pairs exactly as in Excel
    ordered_pairs = [
        "GBPUSD",
        "EURUSD",
        "EURGBP",
        "NZDUSD",
        "AUDNZD",
        "GBPAUD",
        "GBPNZD",
        "USDCHF",
        "USDJPY",
        "USDCAD",
        "GBPJPY"
    ]

    # Make a request to the API
    response = requests.get(f"{api_url}?apikey={api_key}")

    # Check if the request was successful
    if response.status_code == 200:
        data = response.json()
        data_dict = {item['symbol']: item for item in data}
        
        # Print the filtered data
        print("\nCurrency Pivot Points and Levels:")
        print("=" * 80)
        
        # Process pairs in the exact order
        for symbol in ordered_pairs:
            if symbol in data_dict:
                item = data_dict[symbol]
                currency = Currency(
                    symbol=item['symbol'],
                    high=item['dayHigh'],
                    low=item['dayLow'],
                    close=item['price']
                )
                currencies.append(currency)
                print(currency)
                print("-" * 80)
            
        update_excel(currencies)
    else:
        print("Failed to retrieve data from the API")

def update_excel(currencies, filename: str = "/Users/eyupensaralmaz/Desktop/twoblok.xlsx"):
    try:
        # Load existing workbook
        book = load_workbook(filename)
        sheet = book.active
        
        # Write data to specific cells (B7:E17)
        for i, currency in enumerate(currencies, start=7):
            # Support Level 1 (B column)
            sheet.cell(row=i, column=2, value=round(currency.support_1, 5))
            
            # Support Level 2 (C column)
            sheet.cell(row=i, column=3, value=round(currency.support_2, 5))
            
            # Resistance Level 1 (D column)
            sheet.cell(row=i, column=4, value=round(currency.resistance_1, 5))
            
            # Resistance Level 2 (E column)
            sheet.cell(row=i, column=5, value=round(currency.resistance_2, 5))
        
        # Save the workbook
        book.save(filename)
        print(f"\nCurrency data successfully updated in {filename}")
        print("Updated cells: B7-B17 (S1), C7-C17 (S2), D7-D17 (R1), E7-E17 (R2)")
        
    except Exception as e:
        print(f"Error updating Excel file: {e}")
