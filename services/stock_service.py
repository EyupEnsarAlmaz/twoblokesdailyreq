import requests
from typing import List
from model.stock_model import Stock
from openpyxl import load_workbook
from datetime import datetime


class StockAPI:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://financialmodelingprep.com/api/v3"

    @staticmethod
    def get_day_suffix(day):
        """Return the suffix for the day (st, nd, rd, th)"""
        if 10 <= day % 100 <= 20:
            suffix = 'th'
        else:
            suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
        return suffix

    def get_formatted_date(self):
        now = datetime.now()
        day = now.day
        suffix = self.get_day_suffix(day)
        return now.strftime(f"%-d{suffix} %B %Y")

    def get_stock_quotes(self, symbols: List[str]) -> List[Stock]:
        url = f"{self.base_url}/quote/{','.join(symbols)}"
        params = {"apikey": self.api_key}

        try:
            response = requests.get(url, params=params)
            response.raise_for_status()
            data = response.json()
            return [Stock.from_dict(item) for item in data]
        except requests.exceptions.RequestException as e:
            print(f"Error fetching data: {e}")
            return []

    def update_excel(self, stocks: List[Stock], filename: str = "/Users/eyupensaralmaz/Desktop/twoblokes.xlsx"):
        """Update existing Excel file with stock data"""
        try:
            # Load existing workbook
            book = load_workbook(filename)
            
            # Get active sheet
            sheet = book.active
            
            # Use formatted date with suffix
            formatted_date = self.get_formatted_date()
            sheet['A2'] = formatted_date
            
            # Write data to specific cells (B7:E16)
            for i, stock in enumerate(stocks, start=7):
                # Support Level 1 (B column)
                sheet.cell(row=i, column=2, value=round(stock.support_1, 2))
                
                # Support Level 2 (C column)
                sheet.cell(row=i, column=3, value=round(stock.support_2, 2))
                
                # Resistance Level 1 (D column)
                sheet.cell(row=i, column=4, value=round(stock.resistance_1, 2))
                
                # Resistance Level 2 (E column)
                sheet.cell(row=i, column=5, value=round(stock.resistance_2, 2))
            
            # Save the workbook
            book.save(filename)
            print(f"\nData successfully updated in {filename}")
            print("Updated cells: A2 (Date), B7-B16 (S1), C7-C16 (S2), D7-D16 (R1), E7-E16 (R2)")
            
        except Exception as e:
            print(f"Error updating Excel file: {e}")

